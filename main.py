from selenium import webdriver
from selenium.webdriver.chrome.service import Service
import time 
from openpyxl import load_workbook

#Variables
URL = "https://rpachallenge.com"
START_BUTTON = "//button"
FIRST_NAME_FORM = "//input[@ng-reflect-name='labelFirstName']"
LAST_NAME_FORM = "//input[@ng-reflect-name='labelLastName']"
PHONE_FORM = "//input[@ng-reflect-name='labelPhone']"
EMAIL_FORM = "//input[@ng-reflect-name='labelEmail']"
ADDRESS_FORM = "//input[@ng-reflect-name='labelAddress']"
COMPANY_FORM = "//input[@ng-reflect-name='labelCompanyName']"
ROLE_FORM = "//input[@ng-reflect-name='labelRole']"
SUBMIT_BUTTON = "//input[@type='submit']"
SUCCESS = "//div[@class='congratulations col s8 m8 l8']"


def main():
    website = URL
    #You need to change this to your own path
    path = "/Users/cindyz/Applications/chromedriver-mac-x64/chromedriver"

    # Default Service instance
    service = Service(executable_path=path)
    driver = webdriver.Chrome(service=service)

    # Navigate to a page given by the URL
    driver.get(website)


    wb = load_workbook("challenge.xlsx")
    sheet = wb.active


    start = driver.find_element(by = "xpath", value = START_BUTTON).click()

    for i in range(2 , 12):
        first_name = driver.find_element(by = "xpath", value = FIRST_NAME_FORM).send_keys(sheet.cell(i, 1).value)
        last_name = driver.find_element(by = "xpath", value = LAST_NAME_FORM).send_keys(sheet.cell(i, 2).value)
        company = driver.find_element(by = "xpath", value = COMPANY_FORM).send_keys(sheet.cell(i, 3).value)
        role = driver.find_element(by = "xpath", value = ROLE_FORM).send_keys(sheet.cell(i, 4).value)
        address = driver.find_element(by = "xpath", value = ADDRESS_FORM).send_keys(sheet.cell(i, 5).value)
        email = driver.find_element(by = "xpath", value = EMAIL_FORM).send_keys(sheet.cell(i, 6).value)
        phone = driver.find_element(by = "xpath", value = PHONE_FORM).send_keys(sheet.cell(i, 7).value)
        submit = driver.find_element(by = "xpath", value = SUBMIT_BUTTON).click()

    driver.save_screenshot("result.png")
    time.sleep(10)

if __name__ == "__main__":
    main()
